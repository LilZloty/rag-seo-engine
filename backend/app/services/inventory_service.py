"""
Inventory Service — Core inventory sync, analysis, and intelligence engine.
Syncs inventory data from Shopify via REST API, processes webhooks,
calculates stock health metrics, and generates inventory alerts.
"""
import uuid
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func as sql_func

from app.models.product import Product
from app.models.inventory_models import InventorySnapshot, InventoryAlert, RestockEvent, OrderLineItem
from app.services.redis_service import cache


class InventoryService:
    """Core inventory intelligence engine."""

    def __init__(self, db: Session):
        self.db = db

    # =========================================================================
    # SHOPIFY SYNC
    # =========================================================================

    def sync_inventory_from_shopify(self) -> Dict[str, Any]:
        """
        Full inventory sync via Shopify REST API.
        Uses the shopify Python library to fetch all products + variants,
        then updates Product.inventory_quantity and inventory_status.
        Returns sync stats.
        """
        import shopify as shopify_lib
        from app.services.shopify_service import ShopifyService
        shopify_svc = ShopifyService()

        if not shopify_svc._ensure_initialized():
            return {"error": "Shopify not initialized", "products_synced": 0}

        # Fetch all products with variants via REST API (paginated)
        all_shopify_products = []
        page_num = 1
        page = shopify_lib.Product.find(limit=250, fields="id,variants")

        while page:
            all_shopify_products.extend(page)
            print(f"[Inventory] Fetched page {page_num}: {len(all_shopify_products)} products so far")
            if page.has_next_page():
                page = page.next_page()
                page_num += 1
            else:
                break

        # Fetch per-location inventory via GraphQL (doesn't need read_locations scope)
        product_location_inventory: Dict[str, Dict[str, int]] = {}
        try:
            product_location_inventory = self._fetch_location_inventory_graphql(shopify_svc)
        except Exception as e:
            print(f"[Inventory] Location-level fetch failed: {e}")

        # Aggregate inventory by product (sum variant quantities)
        product_inventory: Dict[str, int] = {}
        for sp in all_shopify_products:
            shopify_id = str(sp.id)
            total_qty = 0
            if hasattr(sp, 'variants') and sp.variants:
                for variant in sp.variants:
                    qty = getattr(variant, 'inventory_quantity', 0)
                    total_qty += (qty if qty and qty > 0 else 0)
            product_inventory[shopify_id] = total_qty

        print(f"[Inventory] Total products with inventory data: {len(product_inventory)}")

        # Update products in database
        now = datetime.now(timezone.utc)
        synced = 0
        qty_changed = 0
        oos_count = 0
        low_stock_count = 0
        restock_events = []
        waitlist_cleared = []  # Products restocked that had waitlist subscribers

        products = self.db.query(Product).all()
        for product in products:
            shopify_id = product.shopify_id
            if shopify_id not in product_inventory:
                continue

            new_qty = product_inventory[shopify_id]
            old_qty = product.inventory_quantity or 0
            old_status = product.inventory_status
            threshold = product.low_stock_threshold or 5

            if new_qty != old_qty:
                qty_changed += 1

            # Determine new status
            if new_qty <= 0:
                new_status = "out_of_stock"
                oos_count += 1
            elif new_qty <= threshold:
                new_status = "low_stock"
                low_stock_count += 1
            else:
                new_status = "in_stock"

            # Detect restock event (was OOS, now has stock)
            if old_status == "out_of_stock" and new_qty > 0:
                restock_event = RestockEvent(
                    id=str(uuid.uuid4()),
                    product_id=product.id,
                    previous_quantity=old_qty,
                    new_quantity=new_qty,
                    restocked_at=now,
                )
                self.db.add(restock_event)
                restock_events.append(product.title)

                # Create back_in_stock alert
                self._create_alert(
                    product, "back_in_stock", "info",
                    f"Back in stock: {product.title} ({new_qty} units)"
                )

                # Clear waitlist subscribers — product is back in stock
                if product.active_subscribers and product.active_subscribers > 0:
                    waitlist_cleared.append({
                        "title": product.title,
                        "sku": product.sku or "",
                        "subscribers_cleared": product.active_subscribers,
                        "new_quantity": new_qty,
                    })
                    product.active_subscribers = 0
                    product.demand_score = self._calculate_demand_score(product)
                    product.urgency_score = self._calculate_urgency_score(product)

            # Detect new stockout
            if old_status != "out_of_stock" and new_qty <= 0:
                product.last_stockout_date = now
                product.stockout_frequency_90d = (product.stockout_frequency_90d or 0) + 1
                self._create_alert(
                    product, "out_of_stock", "critical",
                    f"Out of stock: {product.title}"
                )

            # Detect low stock
            if new_status == "low_stock" and old_status != "low_stock":
                self._create_alert(
                    product, "low_stock", "warning",
                    f"Low stock: {product.title} ({new_qty} units, threshold: {threshold})"
                )

            # Update product
            product.inventory_quantity = new_qty
            product.inventory_status = new_status
            product.last_inventory_sync = now

            # Store per-location breakdown
            if shopify_id in product_location_inventory:
                product.inventory_by_location = product_location_inventory[shopify_id]

            # Recalculate velocity and health
            self._update_computed_fields(product)

            synced += 1

        self.db.commit()

        # Invalidate cache
        cache.invalidate("inventory:dashboard")
        cache.invalidate("inventory:health")

        stats = {
            "products_synced": synced,
            "qty_changed": qty_changed,
            "total_variants_fetched": sum(
                len(sp.variants) if hasattr(sp, 'variants') and sp.variants else 0
                for sp in all_shopify_products
            ),
            "out_of_stock": oos_count,
            "low_stock": low_stock_count,
            "restock_events": restock_events,
            "waitlist_cleared": waitlist_cleared,
            "waitlist_cleared_count": len(waitlist_cleared),
            "synced_at": now.isoformat(),
        }
        print(f"[Inventory] Sync complete: {stats}")
        return stats

    # =========================================================================
    # WEBHOOK PROCESSING
    # =========================================================================

    def process_inventory_webhook(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """
        Handle Shopify inventory_levels/update webhook.
        Payload contains: inventory_item_id, location_id, available, updated_at
        """
        inventory_item_id = str(payload.get("inventory_item_id", ""))
        available = payload.get("available", 0)

        # Find the product by matching inventory_item_id
        # We need to look up via Shopify API since we store shopify_id not inventory_item_id
        # For now, log and trigger a targeted sync
        print(f"[Inventory Webhook] Item {inventory_item_id} -> available: {available}")

        # Create a snapshot record
        # Note: In Phase 2, we'll map inventory_item_id to product_id properly
        return {
            "processed": True,
            "inventory_item_id": inventory_item_id,
            "available": available,
        }

    # =========================================================================
    # COMPUTED FIELDS
    # =========================================================================

    def _update_computed_fields(self, product: Product):
        """Recalculate velocity, days of supply, stock health, demand score, urgency, and reorder qty."""
        import math

        # Velocity: units sold per day (30d average)
        sold_30d = product.sold_30d or 0
        sold_90d = product.sold_90d or 0
        sold_365d = product.sold_365d or 0
        velocity = sold_30d / 30.0 if sold_30d > 0 else 0.0
        product.inventory_velocity = round(velocity, 3)

        # Days of supply
        qty = product.inventory_quantity or 0
        if velocity > 0 and qty > 0:
            product.days_of_supply = round(qty / velocity, 1)
        elif qty > 0:
            product.days_of_supply = 999.0  # No velocity = infinite supply
        else:
            product.days_of_supply = 0.0  # No stock

        # Stock health (uses 90d window for "dead" — products with NO sales in last 90 days)
        threshold = product.low_stock_threshold or 5
        if qty <= 0:
            product.stock_health = "critical"
        elif qty <= threshold:
            product.stock_health = "warning"
        elif qty > 0 and sold_90d == 0:
            # Has stock but no sales in 90 days = dead stock
            product.stock_health = "dead"
        else:
            product.stock_health = "healthy"

        # Dead stock tier (granular sub-classification)
        product.dead_stock_tier = self._classify_dead_stock_tier(
            qty=qty, sold_30d=sold_30d, sold_90d=sold_90d, sold_365d=sold_365d, velocity=velocity
        )

        # Demand score (0-100)
        product.demand_score = self._calculate_demand_score(product)

        # --- Action Center fields ---

        # Urgency score
        product.urgency_score = self._calculate_urgency_score(product)

        # Revenue lost estimate (OOS products only)
        product.revenue_lost_est = self._calculate_revenue_lost(product)

        # Suggested reorder quantity
        product.suggested_reorder_qty = self._calculate_suggested_reorder(product)

    def _classify_dead_stock_tier(
        self,
        qty: int,
        sold_30d: int,
        sold_90d: int,
        sold_365d: int,
        velocity: float,
    ) -> Optional[str]:
        """
        Classify a product into a dead-stock tier based on sales recency.

        Tiers (only set when product has stock):
          - "slow"     : Sales in last 30d but velocity below 0.05/day (~1.5 units/month)
          - "stale"    : 0 sales in last 30d, but had sales in 31-90d window
          - "dead"     : 0 sales in last 90d, but had sales in 91-365d window
          - "obsolete" : 0 sales in last 365d (zombie inventory)
          - None       : Healthy seller (no flag)

        OOS products (qty <= 0) are excluded — they have nothing to flag.
        """
        if qty <= 0:
            return None

        if sold_365d == 0:
            return "obsolete"
        if sold_90d == 0:
            return "dead"
        if sold_30d == 0:
            return "stale"
        # Has recent sales — but is it slow?
        # Threshold: less than ~1.5 units/month (0.05/day)
        if velocity > 0 and velocity < 0.05:
            return "slow"
        return None

    def _calculate_demand_score(self, product: Product) -> int:
        """
        Composite demand score (0-100).

        Base score (0-100) from 6 universally available signals:
          - Velocity 25%: units/day from actual sales
          - Revenue 20%: revenue generated in 30d
          - Sales trend 20%: acceleration vs deceleration (30d vs 90d avg)
          - GA4 sessions 15%: people visiting the product page
          - GSC impressions 10%: search visibility (people searching for it)
          - GSC clicks 10%: confirmed search demand (people clicking through)

        Bonus multipliers (applied AFTER base score):
          - Subscribers: if waitlist > 0, boost score by up to 20%
          - Anchor score: if product drives multi-product carts, boost by up to 15%

        This ensures products without waitlist/co-purchase data can still score 100.
        """
        velocity = product.inventory_velocity or 0.0
        revenue_30d = product.revenue_30d or 0.0
        sessions = product.ga4_sessions or 0
        gsc_impressions = product.gsc_impressions or 0
        gsc_clicks = product.gsc_clicks or 0
        sold_30d = product.sold_30d or 0
        sold_90d = product.sold_90d or 0

        # Sales trend: compare last 30d pace vs 90d avg pace
        avg_monthly_90d = sold_90d / 3.0 if sold_90d > 0 else 0
        if avg_monthly_90d > 0:
            trend_ratio = sold_30d / avg_monthly_90d
        elif sold_30d > 0:
            trend_ratio = 2.0  # New demand (had no 90d history)
        else:
            trend_ratio = 0.0

        # Base score from universal signals (any product can reach 100)
        base_score = (
            min(velocity / 5.0, 1.0) * 25 +              # 25% sales velocity
            min(revenue_30d / 5000.0, 1.0) * 20 +         # 20% revenue
            min(trend_ratio / 2.0, 1.0) * 20 +            # 20% demand trend
            min(sessions / 200.0, 1.0) * 15 +             # 15% GA4 traffic
            min(gsc_impressions / 500.0, 1.0) * 10 +      # 10% search visibility
            min(gsc_clicks / 50.0, 1.0) * 10              # 10% search clicks
        )

        # Bonus: Waitlist subscribers (boost, not penalty)
        subscribers = product.active_subscribers or 0
        if subscribers > 0:
            subscriber_boost = min(subscribers / 50.0, 1.0) * 20  # up to +20 points
            base_score += subscriber_boost

        # Bonus: Anchor product (drives multi-product carts)
        anchor = product.anchor_score or 0
        if anchor > 0:
            anchor_boost = (anchor / 100.0) * 15  # up to +15 points
            base_score += anchor_boost

        return min(int(round(base_score)), 100)

    def _calculate_urgency_score(self, product: Product) -> int:
        """
        Urgency score for Action Center row coloring.
        Returns -1 for dead stock (any non-slow tier), 0-100 for everything else.
        >=70 = critical (red), >=40 = warning (yellow), >=1 = healthy (green).
        """
        import math

        qty = product.inventory_quantity or 0
        velocity = product.inventory_velocity or 0.0
        sold_30d = product.sold_30d or 0
        sold_90d = product.sold_90d or 0
        demand = product.demand_score or 0
        days_supply = product.days_of_supply or 0.0
        stockout_freq = product.stockout_frequency_90d or 0

        # Dead stock: special flag (uses 90d window — no sales in last 90 days)
        if qty > 0 and sold_90d == 0:
            return -1

        urgency = 0

        # OOS products
        if product.inventory_status == "out_of_stock":
            urgency += 50                                # Base: it's OOS
            urgency += demand * 0.3                      # Higher demand = more urgent
            urgency += min(sold_30d * 2, 20)             # Was selling? More urgent
            urgency += min(stockout_freq * 5, 15)        # Repeat offender bonus

        # Low stock products
        elif product.inventory_status == "low_stock":
            urgency += 30                                # Base: it's low
            if days_supply <= 3:
                urgency += 30                            # Critical: 3 days or less
            elif days_supply <= 7:
                urgency += 15
            elif days_supply <= 14:
                urgency += 5

        # In-stock but running out fast
        elif product.inventory_status == "in_stock" and days_supply > 0:
            if days_supply <= 3:
                urgency += 50                            # About to run out
            elif days_supply <= 7:
                urgency += 25

        return min(int(round(urgency)), 100)

    def _calculate_revenue_lost(self, product: Product) -> float:
        """
        Estimated revenue lost while OOS.
        Formula: min(days_since_stockout, 30) * daily_velocity * price
        """
        if product.inventory_status != "out_of_stock":
            return 0.0

        velocity = product.inventory_velocity or 0.0
        if velocity <= 0:
            return 0.0

        try:
            price = float(product.price) if product.price else 0.0
        except (ValueError, TypeError):
            price = 0.0

        if price <= 0:
            return 0.0

        # Days since stockout
        days_oos = 30  # default max
        if product.last_stockout_date:
            delta = datetime.now(timezone.utc) - product.last_stockout_date
            days_oos = min(delta.days, 30)
            days_oos = max(days_oos, 1)  # at least 1 day

        return round(days_oos * velocity * price, 2)

    def _calculate_suggested_reorder(self, product: Product) -> int:
        """
        Suggested reorder quantity to cover 30 days of demand + 20% safety buffer.
        """
        import math

        velocity = product.inventory_velocity or 0.0
        if velocity <= 0:
            return 0

        qty = product.inventory_quantity or 0
        target_days = 30
        safety_factor = 1.2
        needed = math.ceil(velocity * target_days * safety_factor) - qty
        return max(needed, 0)

    # =========================================================================
    # LOCATION-LEVEL INVENTORY (GraphQL)
    # =========================================================================

    def _fetch_location_inventory_graphql(self, shopify_svc) -> Dict[str, Dict[str, int]]:
        """
        Fetch per-location inventory for all products via Shopify GraphQL.
        Uses only read_inventory scope (location.id only, not name).
        Then resolves location names via a separate fulfillmentService-free query.
        Returns: {shopify_product_id: {"Location Name": qty, ...}}
        """
        import requests

        from app.core.config import settings
        url = f"https://{settings.SHOPIFY_STORE}/admin/api/{settings.SHOPIFY_API_VERSION}/graphql.json"
        headers = {
            "Content-Type": "application/json",
            "X-Shopify-Access-Token": settings.SHOPIFY_ACCESS_TOKEN,
        }

        # Step 1: Try to get location names via fulfillmentOrders or shop query
        location_names: Dict[str, str] = {}  # gid -> name

        # Try locations query with read_locations scope first
        loc_query = "{ locations(first: 50) { edges { node { id name } } } }"
        try:
            resp = requests.post(url, json={"query": loc_query}, headers=headers, timeout=10)
            data = resp.json()
            if "errors" not in data:
                for edge in data.get("data", {}).get("locations", {}).get("edges", []):
                    loc_id = edge["node"]["id"]
                    location_names[loc_id] = edge["node"]["name"]
                print(f"[Inventory] Resolved {len(location_names)} location names via locations query")
        except Exception:
            pass

        # Fallback: known Example Store location mapping
        if not location_names:
            location_names = {
                "gid://shopify/Location/1111111111": "Main Warehouse",
                "gid://shopify/Location/2222222222": "Secondary Warehouse",
            }
            print(f"[Inventory] Using known location mapping: {list(location_names.values())}")

        # Step 2: Fetch inventory levels per product (no location.name, only location.id)
        result: Dict[str, Dict[str, int]] = {}
        cursor = None
        page_count = 0

        query = """
        query($cursor: String) {
          products(first: 50, after: $cursor) {
            pageInfo { hasNextPage endCursor }
            edges {
              node {
                id
                variants(first: 20) {
                  edges {
                    node {
                      inventoryItem {
                        inventoryLevels(first: 20) {
                          edges {
                            node {
                              quantities(names: ["available"]) {
                                quantity
                              }
                              location {
                                id
                              }
                            }
                          }
                        }
                      }
                    }
                  }
                }
              }
            }
          }
        }
        """

        while True:
            page_count += 1
            variables = {"cursor": cursor} if cursor else {}
            resp = requests.post(url, json={"query": query, "variables": variables}, headers=headers, timeout=30)
            data = resp.json()

            if "errors" in data:
                print(f"[Inventory] GraphQL inventory error: {data['errors']}")
                break

            products_data = data.get("data", {}).get("products", {})
            edges = products_data.get("edges", [])

            for product_edge in edges:
                node = product_edge["node"]
                shopify_id = node["id"].split("/")[-1]

                location_qtys: Dict[str, int] = {}
                for variant_edge in node.get("variants", {}).get("edges", []):
                    inv_item = variant_edge["node"].get("inventoryItem", {})
                    for level_edge in inv_item.get("inventoryLevels", {}).get("edges", []):
                        level_node = level_edge["node"]
                        loc_gid = level_node.get("location", {}).get("id", "")
                        loc_label = location_names.get(loc_gid, f"Sucursal {loc_gid.split('/')[-1]}")
                        quantities = level_node.get("quantities", [])
                        qty = quantities[0]["quantity"] if quantities else 0
                        location_qtys[loc_label] = location_qtys.get(loc_label, 0) + max(qty, 0)

                if location_qtys:
                    result[shopify_id] = location_qtys

            page_info = products_data.get("pageInfo", {})
            if page_info.get("hasNextPage"):
                cursor = page_info["endCursor"]
                if page_count % 10 == 0:
                    print(f"[Inventory] Location fetch page {page_count}: {len(result)} products so far")
            else:
                break

        print(f"[Inventory] Location-level data fetched for {len(result)} products ({page_count} pages)")
        return result

    # =========================================================================
    # ALERTS
    # =========================================================================

    def _create_alert(self, product: Product, alert_type: str, severity: str, message: str):
        """Create an inventory alert, avoiding duplicates."""
        # Check if an active alert of this type already exists for this product
        existing = self.db.query(InventoryAlert).filter(
            InventoryAlert.product_id == product.id,
            InventoryAlert.alert_type == alert_type,
            InventoryAlert.status == "active",
        ).first()

        if existing:
            return  # Don't duplicate

        alert = InventoryAlert(
            id=str(uuid.uuid4()),
            alert_type=alert_type,
            product_id=product.id,
            message=message,
            severity=severity,
            status="active",
        )
        self.db.add(alert)

    def get_alerts(self, status: str = "active", limit: int = 50) -> List[Dict]:
        """Get inventory alerts."""
        query = self.db.query(InventoryAlert).filter(
            InventoryAlert.status == status
        ).order_by(InventoryAlert.triggered_at.desc()).limit(limit)

        return [
            {
                "id": a.id,
                "type": a.alert_type,
                "product_id": a.product_id,
                "message": a.message,
                "severity": a.severity,
                "status": a.status,
                "triggered_at": a.triggered_at.isoformat() if a.triggered_at else None,
            }
            for a in query.all()
        ]

    def acknowledge_alert(self, alert_id: str) -> bool:
        """Mark an alert as acknowledged."""
        alert = self.db.query(InventoryAlert).filter(InventoryAlert.id == alert_id).first()
        if not alert:
            return False
        alert.status = "acknowledged"
        alert.acknowledged_at = datetime.now(timezone.utc)
        self.db.commit()
        return True

    # =========================================================================
    # DASHBOARD & ANALYTICS
    # =========================================================================

    def get_inventory_dashboard(self) -> Dict[str, Any]:
        """Aggregated inventory overview for the dashboard."""
        # Check cache
        cached = cache.get("inventory:dashboard")
        if cached:
            return cached

        products = self.db.query(Product).all()
        total = len(products)

        in_stock = sum(1 for p in products if p.inventory_status == "in_stock")
        out_of_stock = sum(1 for p in products if p.inventory_status == "out_of_stock")
        low_stock = sum(1 for p in products if p.inventory_status == "low_stock")
        not_synced = sum(1 for p in products if p.inventory_status is None)

        # Dead stock: has inventory but no sales in 90d (matches stock_health="dead")
        dead_stock = sum(
            1 for p in products
            if p.stock_health == "dead"
        )

        # Tier breakdown — granular sub-classification of slow/inactive products
        tier_counts = {"slow": 0, "stale": 0, "dead": 0, "obsolete": 0}
        for p in products:
            tier = p.dead_stock_tier
            if tier in tier_counts:
                tier_counts[tier] += 1

        # Active alerts count
        active_alerts = self.db.query(InventoryAlert).filter(
            InventoryAlert.status == "active"
        ).count()

        # Recent restocks (last 7 days)
        week_ago = datetime.now(timezone.utc) - timedelta(days=7)
        recent_restocks = self.db.query(RestockEvent).filter(
            RestockEvent.restocked_at >= week_ago
        ).count()

        # Average days of supply (for in-stock products that have velocity)
        in_stock_with_velocity = [
            p for p in products
            if p.days_of_supply and p.days_of_supply < 999 and p.inventory_status == "in_stock"
        ]
        avg_days_supply = (
            sum(p.days_of_supply for p in in_stock_with_velocity) / len(in_stock_with_velocity)
            if in_stock_with_velocity else 0
        )

        result = {
            "total_products": total,
            "in_stock": in_stock,
            "out_of_stock": out_of_stock,
            "low_stock": low_stock,
            "not_synced": not_synced,
            "dead_stock": dead_stock,
            "dead_stock_tiers": tier_counts,  # {slow, stale, dead, obsolete}
            "in_stock_rate": round(in_stock / total * 100, 1) if total > 0 else 0,
            "active_alerts": active_alerts,
            "recent_restocks_7d": recent_restocks,
            "avg_days_of_supply": round(avg_days_supply, 1),
            "last_sync": max(
                (p.last_inventory_sync for p in products if p.last_inventory_sync),
                default=None
            ),
        }

        cache.set("inventory:dashboard", result, ttl=300)
        return result

    def get_products_by_status(
        self, status: Optional[str] = None, sort_by: str = "demand_score", limit: int = 100
    ) -> List[Dict]:
        """Get products with inventory data, filterable by status. Used by other endpoints."""
        query = self.db.query(Product)
        if status:
            query = query.filter(Product.inventory_status == status)
        query = self._apply_sort(query, sort_by)
        products = query.limit(limit).all()
        return [self._product_to_inventory_dict(p) for p in products]

    def get_products_paginated(
        self,
        status: Optional[str] = None,
        sort_by: str = "demand_score",
        search: Optional[str] = None,
        page: int = 1,
        page_size: int = 50,
    ) -> Dict[str, Any]:
        """Server-side paginated product list with filtering and search."""
        from sqlalchemy import or_

        query = self.db.query(Product).filter(Product.inventory_status.isnot(None))

        # Status filter
        if status == "dead_stock":
            query = query.filter(Product.stock_health == "dead", Product.inventory_quantity > 0)
        elif status in ("slow", "stale", "dead", "obsolete"):
            # Granular tier filter (sub-classification of dead/inactive)
            query = query.filter(Product.dead_stock_tier == status, Product.inventory_quantity > 0)
        elif status:
            query = query.filter(Product.inventory_status == status)

        # Search
        if search and search.strip():
            term = f"%{search.strip()}%"
            query = query.filter(
                or_(
                    Product.title.ilike(term),
                    Product.sku.ilike(term),
                    Product.handle.ilike(term),
                    Product.product_type.ilike(term),
                    Product.shopify_id.ilike(term),
                )
            )

        # Get total count before pagination
        total = query.count()

        # Sort
        query = self._apply_sort(query, sort_by)

        # Paginate
        offset = (page - 1) * page_size
        products = query.offset(offset).limit(page_size).all()

        # Status counts (cached for the full filtered set)
        count_query = self.db.query(Product).filter(Product.inventory_status.isnot(None))
        if search and search.strip():
            term = f"%{search.strip()}%"
            count_query = count_query.filter(
                or_(
                    Product.title.ilike(term),
                    Product.sku.ilike(term),
                    Product.handle.ilike(term),
                    Product.product_type.ilike(term),
                    Product.shopify_id.ilike(term),
                )
            )

        all_count = count_query.count()
        in_stock_count = count_query.filter(Product.inventory_status == "in_stock").count()
        oos_count = count_query.filter(Product.inventory_status == "out_of_stock").count()
        low_count = count_query.filter(Product.inventory_status == "low_stock").count()
        dead_count = count_query.filter(Product.stock_health == "dead", Product.inventory_quantity > 0).count()

        # Tier counts (granular sub-classification of slow/inactive)
        tier_q = count_query.filter(Product.inventory_quantity > 0)
        slow_count = tier_q.filter(Product.dead_stock_tier == "slow").count()
        stale_count = tier_q.filter(Product.dead_stock_tier == "stale").count()
        tier_dead_count = tier_q.filter(Product.dead_stock_tier == "dead").count()
        obsolete_count = tier_q.filter(Product.dead_stock_tier == "obsolete").count()

        return {
            "products": [self._product_to_inventory_dict(p) for p in products],
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": -(-total // page_size),  # ceil division
            },
            "counts": {
                "all": all_count,
                "in_stock": in_stock_count,
                "out_of_stock": oos_count,
                "low_stock": low_count,
                "dead_stock": dead_count,
                "slow": slow_count,
                "stale": stale_count,
                "dead": tier_dead_count,
                "obsolete": obsolete_count,
            },
        }

    def _apply_sort(self, query, sort_by: str):
        """Apply sort to a product query."""
        if sort_by == "demand_score" or sort_by == "demand":
            return query.order_by(Product.demand_score.desc())
        elif sort_by == "velocity":
            return query.order_by(Product.inventory_velocity.desc().nullslast())
        elif sort_by == "days_of_supply" or sort_by == "supply":
            return query.order_by(Product.days_of_supply.asc().nullslast())
        elif sort_by == "quantity" or sort_by == "qty_asc":
            return query.order_by(Product.inventory_quantity.asc().nullslast())
        elif sort_by == "qty_desc":
            return query.order_by(Product.inventory_quantity.desc().nullslast())
        elif sort_by == "urgency":
            return query.order_by(Product.urgency_score.desc())
        elif sort_by == "title":
            return query.order_by(Product.title.asc())
        elif sort_by == "sku":
            return query.order_by(Product.sku.asc().nullslast())
        elif sort_by == "price_asc":
            return query.order_by(Product.price.asc().nullslast())
        elif sort_by == "price_desc":
            return query.order_by(Product.price.desc().nullslast())
        else:
            return query.order_by(Product.demand_score.desc())

    def get_stockout_risk_products(self, days: int = 7) -> List[Dict]:
        """Products that will run out within N days at current velocity."""
        products = self.db.query(Product).filter(
            Product.inventory_status == "in_stock",
            Product.days_of_supply.isnot(None),
            Product.days_of_supply <= days,
            Product.days_of_supply > 0,
        ).order_by(Product.days_of_supply.asc()).all()

        return [self._product_to_inventory_dict(p) for p in products]

    def get_dead_stock(
        self,
        no_sales_days: int = 90,
        tier: Optional[str] = None,
    ) -> List[Dict]:
        """
        Products with stock but no sales in the requested window.

        Args:
            no_sales_days: Window for "no sales" — uses sold_30d/sold_90d/sold_365d buckets.
                           30 → no sales in 30d, 90 → no sales in 90d (default), 365 → no sales in 365d.
            tier: Optional granular filter — "slow", "stale", "dead", or "obsolete".
                  When set, no_sales_days is ignored (the tier already encodes the window).
        """
        query = self.db.query(Product).filter(Product.inventory_quantity > 0)

        if tier in ("slow", "stale", "dead", "obsolete"):
            query = query.filter(Product.dead_stock_tier == tier)
        elif no_sales_days >= 365:
            query = query.filter((Product.sold_365d == 0) | (Product.sold_365d.is_(None)))
        elif no_sales_days >= 90:
            query = query.filter((Product.sold_90d == 0) | (Product.sold_90d.is_(None)))
        else:
            query = query.filter((Product.sold_30d == 0) | (Product.sold_30d.is_(None)))

        products = query.order_by(Product.inventory_quantity.desc()).all()

        # Enrich with capital_tied_up so the frontend can sort by $ at risk
        result = []
        for p in products:
            d = self._product_to_inventory_dict(p)
            try:
                price = float(p.price) if p.price else 0.0
            except (ValueError, TypeError):
                price = 0.0
            d["capital_tied_up"] = round((p.inventory_quantity or 0) * price, 2)
            result.append(d)
        return result

    def get_restock_priority(self) -> List[Dict]:
        """OOS products ranked by demand score (what to restock first)."""
        products = self.db.query(Product).filter(
            Product.inventory_status == "out_of_stock"
        ).order_by(Product.demand_score.desc()).all()

        return [self._product_to_inventory_dict(p) for p in products]

    def get_action_center(self) -> Dict[str, Any]:
        """
        Action Center: grouped product lists by what action to take.
        Returns 4 sections: restock_now, order_soon, slow_movers, star_products.
        """
        cached = cache.get("inventory:action_center")
        if cached:
            return cached

        products = self.db.query(Product).filter(
            Product.inventory_status.isnot(None)
        ).all()

        restock_now = []
        order_soon = []
        slow_movers = []
        star_products = []

        # Calculate median velocity for star product threshold
        velocities = [p.inventory_velocity for p in products if p.inventory_velocity and p.inventory_velocity > 0]
        median_velocity = sorted(velocities)[len(velocities) // 2] if velocities else 0.0

        for p in products:
            d = self._product_to_inventory_dict(p)
            urgency = p.urgency_score or 0
            dos = p.days_of_supply or 0.0
            vel = p.inventory_velocity or 0.0
            qty = p.inventory_quantity or 0

            # Restock Now: urgency >= 70 (critical)
            if urgency >= 70:
                restock_now.append(d)

            # Order Soon: urgency 40-69 (warning) OR in-stock with <= 14 days supply
            elif urgency >= 40 or (p.inventory_status == "in_stock" and 0 < dos <= 14 and vel > 0):
                order_soon.append(d)

            # Slow Movers: dead stock flag (urgency == -1)
            elif urgency == -1:
                try:
                    capital = qty * float(p.price) if p.price else 0.0
                except (ValueError, TypeError):
                    capital = 0.0
                d["capital_tied_up"] = round(capital, 2)
                slow_movers.append(d)

            # Star Products: healthy, good velocity, sustainable supply
            elif (p.stock_health == "healthy" and vel >= median_velocity
                  and vel > 0 and 14 <= dos <= 60):
                star_products.append(d)

        # Sort each section
        restock_now.sort(key=lambda x: x.get("urgency_score", 0), reverse=True)
        order_soon.sort(key=lambda x: x.get("days_of_supply") or 999)
        slow_movers.sort(key=lambda x: x.get("capital_tied_up", 0), reverse=True)
        star_products.sort(key=lambda x: x.get("revenue_30d") or 0, reverse=True)

        result = {
            "restock_now": restock_now[:50],
            "order_soon": order_soon[:50],
            "slow_movers": slow_movers[:50],
            "star_products": star_products[:50],
            "counts": {
                "restock_now": len(restock_now),
                "order_soon": len(order_soon),
                "slow_movers": len(slow_movers),
                "star_products": len(star_products),
            },
        }

        cache.set("inventory:action_center", result, ttl=300)
        return result

    def get_revenue_at_risk(self) -> Dict[str, Any]:
        """Total estimated revenue lost from all OOS products."""
        products = self.db.query(Product).filter(
            Product.inventory_status == "out_of_stock",
            Product.revenue_lost_est > 0,
        ).all()

        total_lost = sum(p.revenue_lost_est or 0 for p in products)
        return {
            "total_revenue_lost": round(total_lost, 2),
            "products_affected": len(products),
        }

    def get_product_snapshots(self, product_id: str, days: int = 30) -> List[Dict]:
        """Historical inventory snapshots for a product."""
        since = datetime.now(timezone.utc) - timedelta(days=days)
        snapshots = self.db.query(InventorySnapshot).filter(
            InventorySnapshot.product_id == product_id,
            InventorySnapshot.snapshot_date >= since,
        ).order_by(InventorySnapshot.snapshot_date.asc()).all()

        return [
            {
                "date": s.snapshot_date.isoformat() if s.snapshot_date else None,
                "quantity": s.quantity,
                "status": s.status,
                "type": s.snapshot_type,
            }
            for s in snapshots
        ]

    # =========================================================================
    # DAILY SNAPSHOT JOB
    # =========================================================================

    def take_daily_snapshot(self) -> int:
        """Create InventorySnapshot rows for all products. Called by scheduler."""
        products = self.db.query(Product).filter(
            Product.inventory_quantity.isnot(None)
        ).all()

        now = datetime.now(timezone.utc)
        count = 0
        for p in products:
            snap = InventorySnapshot(
                id=str(uuid.uuid4()),
                product_id=p.id,
                quantity=p.inventory_quantity or 0,
                status=p.inventory_status,
                snapshot_date=now,
                snapshot_type="auto",
            )
            self.db.add(snap)
            count += 1

        self.db.commit()
        print(f"[Inventory] Daily snapshot: {count} products recorded")
        return count

    # =========================================================================
    # HELPERS
    # =========================================================================

    # =========================================================================
    # WAITLIST (Back-in-Stock Subscribers)
    # =========================================================================

    def get_waitlist_products(self, sort_by: str = "subscribers") -> List[Dict]:
        """
        Products with active back-in-stock subscribers (AMP by Aisle).
        Returns all products where active_subscribers > 0, sorted by subscriber count or demand.
        """
        query = self.db.query(Product).filter(
            Product.active_subscribers > 0
        )

        if sort_by == "demand":
            query = query.order_by(Product.demand_score.desc())
        elif sort_by == "revenue_lost":
            query = query.order_by(Product.revenue_lost_est.desc())
        else:  # default: subscribers
            query = query.order_by(Product.active_subscribers.desc())

        products = query.all()
        result = []
        for p in products:
            d = self._product_to_inventory_dict(p)
            # Add waitlist-specific computed fields
            try:
                price = float(p.price) if p.price else 0.0
            except (ValueError, TypeError):
                price = 0.0
            d["potential_revenue"] = round((p.active_subscribers or 0) * price, 2)
            result.append(d)
        return result

    def get_waitlist_summary(self) -> Dict[str, Any]:
        """Waitlist overview stats."""
        products = self.db.query(Product).filter(
            Product.active_subscribers > 0
        ).all()

        total_subscribers = sum(p.active_subscribers or 0 for p in products)
        oos_with_subs = [p for p in products if p.inventory_status == "out_of_stock"]
        in_stock_with_subs = [p for p in products if p.inventory_status == "in_stock"]

        # Potential revenue from OOS products with subscribers
        potential_revenue = 0.0
        for p in oos_with_subs:
            try:
                price = float(p.price) if p.price else 0.0
            except (ValueError, TypeError):
                price = 0.0
            potential_revenue += (p.active_subscribers or 0) * price

        return {
            "total_products_with_subscribers": len(products),
            "total_subscribers": total_subscribers,
            "oos_with_subscribers": len(oos_with_subs),
            "in_stock_with_subscribers": len(in_stock_with_subs),
            "potential_revenue": round(potential_revenue, 2),
        }

    def import_subscriber_counts(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Import subscriber counts from AMP CSV export or manual input.
        Each item should have: product_id or shopify_id or sku, and subscriber_count.
        Replaces the full waitlist: products not in the new CSV get subscribers reset to 0.
        """
        updated = 0
        not_found = []
        matched_product_ids = set()

        for item in data:
            count = int(item.get("subscriber_count", 0))
            product = None

            # Try to find product by various identifiers
            if item.get("product_id"):
                product = self.db.query(Product).filter(Product.id == item["product_id"]).first()
            if not product and item.get("shopify_id"):
                product = self.db.query(Product).filter(Product.shopify_id == str(item["shopify_id"])).first()
            if not product and item.get("sku"):
                product = self.db.query(Product).filter(Product.sku == item["sku"]).first()
            if not product and item.get("handle"):
                product = self.db.query(Product).filter(Product.handle == item["handle"]).first()

            if product:
                product.active_subscribers = count
                product.demand_score = self._calculate_demand_score(product)
                product.urgency_score = self._calculate_urgency_score(product)
                matched_product_ids.add(product.id)
                updated += 1
            else:
                identifier = item.get("sku") or item.get("shopify_id") or item.get("product_id") or "unknown"
                not_found.append(identifier)

        # Clear subscribers for products that were on the old waitlist but not in the new CSV
        stale_cleared = 0
        stale_query = self.db.query(Product).filter(Product.active_subscribers > 0)
        if matched_product_ids:
            stale_query = stale_query.filter(~Product.id.in_(matched_product_ids))
        old_waitlist = stale_query.all()
        for p in old_waitlist:
            p.active_subscribers = 0
            p.demand_score = self._calculate_demand_score(p)
            p.urgency_score = self._calculate_urgency_score(p)
            stale_cleared += 1

        self.db.commit()
        cache.invalidate("inventory:dashboard")
        cache.invalidate("inventory:action_center")

        return {
            "updated": updated,
            "stale_cleared": stale_cleared,
            "not_found": not_found,
            "not_found_count": len(not_found),
        }

    # =========================================================================
    # ANALYTICS (server-side computed for full catalog)
    # =========================================================================

    def get_fastest_movers(self, limit: int = 10) -> List[Dict]:
        """Top products by velocity across ALL products, not just current page."""
        products = (
            self.db.query(Product)
            .filter(Product.inventory_velocity > 0)
            .order_by(Product.inventory_velocity.desc())
            .limit(limit)
            .all()
        )
        return [
            {
                "id": p.id,
                "title": p.title,
                "sku": p.sku,
                "inventory_quantity": p.inventory_quantity,
                "inventory_velocity": p.inventory_velocity,
                "days_of_supply": p.days_of_supply,
                "demand_score": p.demand_score or 0,
            }
            for p in products
        ]

    def get_stockout_risk_timeline(self) -> Dict[str, Any]:
        """Count products by stockout risk tier across ALL products."""
        in_stock = (
            self.db.query(Product)
            .filter(
                Product.inventory_status == "in_stock",
                Product.days_of_supply.isnot(None),
                Product.days_of_supply > 0,
            )
            .all()
        )
        tiers = {"today": 0, "3_days": 0, "7_days": 0, "14_days": 0}
        for p in in_stock:
            dos = p.days_of_supply or 999
            if dos <= 1:
                tiers["today"] += 1
            if dos <= 3:
                tiers["3_days"] += 1
            if dos <= 7:
                tiers["7_days"] += 1
            if dos <= 14:
                tiers["14_days"] += 1

        oos_count = self.db.query(Product).filter(
            Product.inventory_status == "out_of_stock"
        ).count()

        return {**tiers, "currently_oos": oos_count}

    # =========================================================================
    # SALES DATA SYNC
    # =========================================================================

    def sync_sales_data(self) -> Dict[str, Any]:
        """
        Sync sold/revenue data from Shopify orders.
        Updates sold_30d, revenue_30d, sold_90d, revenue_90d, etc. on all products.
        Then recalculates computed fields (velocity, demand, urgency) with fresh data.
        """
        from app.services.shopify_service import ShopifyService

        shopify_service = ShopifyService()
        sales_data = shopify_service.get_product_sales_all_periods()

        if not sales_data:
            return {"sales_updated": 0}

        # Pre-load all products into a map (avoid N+1 query)
        all_products = self.db.query(Product).all()
        product_map = {p.shopify_id: p for p in all_products if p.shopify_id}

        updated = 0
        for product_shopify_id, periods in sales_data.items():
            product = product_map.get(str(product_shopify_id))
            if product:
                product.total_sold = periods["90d"]["total_sold"]
                product.total_revenue = periods["90d"]["total_revenue"]
                product.sold_30d = periods["30d"]["total_sold"]
                product.revenue_30d = periods["30d"]["total_revenue"]
                product.sold_90d = periods["90d"]["total_sold"]
                product.revenue_90d = periods["90d"]["total_revenue"]
                product.sold_365d = periods["365d"]["total_sold"]
                product.revenue_365d = periods["365d"]["total_revenue"]
                product.sold_all_time = periods["all_time"]["total_sold"]
                product.revenue_all_time = periods["all_time"]["total_revenue"]

                # Capture date of most recent sale
                last_sold_iso = periods.get("last_sold_at")
                if last_sold_iso:
                    try:
                        # Shopify returns ISO 8601 with timezone
                        dt = datetime.fromisoformat(last_sold_iso.replace("Z", "+00:00"))
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        product.last_sold_at = dt
                    except (ValueError, AttributeError):
                        pass

                # Recompute velocity/demand/urgency/tier with fresh sales
                self._update_computed_fields(product)
                updated += 1

        # Products that didn't appear in sales_data have ZERO sales in last 365d.
        # Reset their counters and re-classify (they may now be obsolete).
        synced_ids = {str(k) for k in sales_data.keys()}
        for product in all_products:
            if product.shopify_id and product.shopify_id not in synced_ids:
                if (product.sold_30d or 0) > 0 or (product.sold_90d or 0) > 0 or (product.sold_365d or 0) > 0:
                    product.sold_30d = 0
                    product.revenue_30d = 0.0
                    product.sold_90d = 0
                    product.revenue_90d = 0.0
                    product.sold_365d = 0
                    product.revenue_365d = 0.0
                    self._update_computed_fields(product)

        self.db.commit()
        cache.invalidate("inventory:dashboard")
        cache.invalidate("inventory:health")
        cache.invalidate("inventory:action_center")

        return {"sales_updated": updated}

    # =========================================================================
    # INCREMENTAL ORDER SYNC (Option A — order_line_items table)
    # =========================================================================

    def sync_orders_incremental(self, force_full: bool = False) -> Dict[str, Any]:
        """
        Incremental sync of Shopify order line items into the order_line_items table.

        Strategy:
          1. Find the latest order_updated_at in our table (or use 365d ago if empty / force_full)
          2. Stream orders from Shopify with updated_at > since_timestamp
          3. UPSERT each line item (handles edits, refunds, cancellations)
          4. Identify affected products (DISTINCT shopify_product_id from upserted rows)
          5. Re-aggregate sold_30d/90d/365d ONLY for affected products from the table
          6. Recompute classification (stock_health, dead_stock_tier, urgency, etc.)

        First sync (empty table) does a full 365-day backfill. Subsequent syncs only
        fetch what changed → typical run is 10-30s instead of 5-10 min.

        Args:
            force_full: If True, ignore the watermark and re-fetch all 365 days.

        Returns:
            stats dict with line_items_upserted, orders_processed, products_affected,
            duration_seconds, sync_window_start
        """
        from app.services.shopify_service import ShopifyService
        from sqlalchemy import select, func as sa_func
        from sqlalchemy.dialects.postgresql import insert as pg_insert
        import time

        start_time = time.time()

        # Determine the sync watermark
        if force_full:
            since_iso = None  # → fetch_orders_since defaults to 365d
            print("[Inventory] sync_orders_incremental: FORCE FULL — fetching all 365d")
        else:
            latest = self.db.execute(
                select(sa_func.max(OrderLineItem.order_updated_at))
            ).scalar()
            if latest is None:
                since_iso = None  # First-ever sync — backfill 365d
                print("[Inventory] sync_orders_incremental: empty table — backfilling 365d")
            else:
                # Subtract 1 hour overlap to catch edge cases (clock skew, retries)
                since = latest - timedelta(hours=1)
                since_iso = since.isoformat()
                print(f"[Inventory] sync_orders_incremental: since {since_iso}")

        shopify_service = ShopifyService()
        affected_products: set = set()
        orders_seen: set = set()
        line_items_upserted = 0

        # Stream + batch upsert (1000 line items per transaction)
        BATCH_SIZE = 1000
        batch: List[Dict[str, Any]] = []

        def flush(batch_rows: List[Dict[str, Any]]):
            """Bulk upsert a batch of line items via INSERT ... ON CONFLICT."""
            if not batch_rows:
                return 0
            stmt = pg_insert(OrderLineItem.__table__).values(batch_rows)
            stmt = stmt.on_conflict_do_update(
                index_elements=['id'],
                set_={
                    'order_id': stmt.excluded.order_id,
                    'order_name': stmt.excluded.order_name,
                    'shopify_product_id': stmt.excluded.shopify_product_id,
                    'shopify_variant_id': stmt.excluded.shopify_variant_id,
                    'sku': stmt.excluded.sku,
                    'title': stmt.excluded.title,
                    'quantity': stmt.excluded.quantity,
                    'current_quantity': stmt.excluded.current_quantity,
                    'unit_price': stmt.excluded.unit_price,
                    'revenue': stmt.excluded.revenue,
                    'is_refunded': stmt.excluded.is_refunded,
                    'is_cancelled': stmt.excluded.is_cancelled,
                    'cancelled_at': stmt.excluded.cancelled_at,
                    'order_updated_at': stmt.excluded.order_updated_at,
                    'fetched_at': sa_func.now(),
                }
            )
            self.db.execute(stmt)
            self.db.commit()
            return len(batch_rows)

        for li in shopify_service.fetch_orders_since(since_iso=since_iso):
            row = {
                'id': li['line_item_id'],
                'order_id': li['order_id'],
                'order_name': li['order_name'],
                'shopify_product_id': li['shopify_product_id'],
                'shopify_variant_id': li['shopify_variant_id'],
                'sku': li['sku'],
                'title': li['title'],
                'quantity': li['quantity'],
                'current_quantity': li['current_quantity'],
                'unit_price': li['unit_price'],
                'revenue': li['revenue'],
                'is_refunded': li['is_refunded'],
                'is_cancelled': li['is_cancelled'],
                'cancelled_at': li['cancelled_at'],
                'order_created_at': li['order_created_at'],
                'order_updated_at': li['order_updated_at'],
            }
            batch.append(row)
            affected_products.add(li['shopify_product_id'])
            orders_seen.add(li['order_id'])

            if len(batch) >= BATCH_SIZE:
                line_items_upserted += flush(batch)
                batch = []

        # Final flush
        line_items_upserted += flush(batch)

        # Re-aggregate counters + recompute classification for affected products
        products_recomputed = 0
        if affected_products:
            print(f"[Inventory] Re-aggregating {len(affected_products)} affected products")
            products_recomputed = self._recompute_products_from_line_items(list(affected_products))

        # If first-ever sync (full backfill), also reset products that have NO line items
        # — they truly have zero sales in 365d and should tip into "obsolete"
        if since_iso is None:
            self._reset_products_with_no_line_items()

        cache.invalidate("inventory:dashboard")
        cache.invalidate("inventory:health")
        cache.invalidate("inventory:action_center")

        duration = time.time() - start_time
        return {
            "line_items_upserted": line_items_upserted,
            "orders_processed": len(orders_seen),
            "products_affected": len(affected_products),
            "products_recomputed": products_recomputed,
            "duration_seconds": round(duration, 1),
            "sync_window_start": since_iso,
            "force_full": force_full,
        }

    def _recompute_products_from_line_items(self, shopify_product_ids: List[str]) -> int:
        """
        Re-aggregate sold_30d/90d/365d/all_time + last_sold_at for the given products
        from the order_line_items table, then refresh classification fields.

        Excludes cancelled and refunded line items from aggregates.
        """
        from sqlalchemy import select, func as sa_func, and_, or_

        if not shopify_product_ids:
            return 0

        now = datetime.now(timezone.utc)
        d30 = now - timedelta(days=30)
        d90 = now - timedelta(days=90)
        d365 = now - timedelta(days=365)

        # One round-trip: aggregate everything per product in a single SQL query
        agg_stmt = select(
            OrderLineItem.shopify_product_id,
            sa_func.coalesce(
                sa_func.sum(OrderLineItem.current_quantity).filter(
                    OrderLineItem.order_created_at >= d30
                ), 0
            ).label('sold_30d'),
            sa_func.coalesce(
                sa_func.sum(OrderLineItem.revenue).filter(
                    OrderLineItem.order_created_at >= d30
                ), 0
            ).label('revenue_30d'),
            sa_func.coalesce(
                sa_func.sum(OrderLineItem.current_quantity).filter(
                    OrderLineItem.order_created_at >= d90
                ), 0
            ).label('sold_90d'),
            sa_func.coalesce(
                sa_func.sum(OrderLineItem.revenue).filter(
                    OrderLineItem.order_created_at >= d90
                ), 0
            ).label('revenue_90d'),
            sa_func.coalesce(
                sa_func.sum(OrderLineItem.current_quantity).filter(
                    OrderLineItem.order_created_at >= d365
                ), 0
            ).label('sold_365d'),
            sa_func.coalesce(
                sa_func.sum(OrderLineItem.revenue).filter(
                    OrderLineItem.order_created_at >= d365
                ), 0
            ).label('revenue_365d'),
            sa_func.coalesce(sa_func.sum(OrderLineItem.current_quantity), 0).label('sold_all'),
            sa_func.coalesce(sa_func.sum(OrderLineItem.revenue), 0).label('revenue_all'),
            sa_func.max(OrderLineItem.order_created_at).label('last_sold_at'),
        ).where(
            and_(
                OrderLineItem.shopify_product_id.in_(shopify_product_ids),
                OrderLineItem.is_cancelled.is_(False),
                OrderLineItem.is_refunded.is_(False),
            )
        ).group_by(OrderLineItem.shopify_product_id)

        rows = self.db.execute(agg_stmt).all()
        agg_by_product = {row.shopify_product_id: row for row in rows}

        # Update each product
        products = self.db.query(Product).filter(
            Product.shopify_id.in_(shopify_product_ids)
        ).all()

        recomputed = 0
        for product in products:
            agg = agg_by_product.get(product.shopify_id)
            if agg:
                product.sold_30d = int(agg.sold_30d or 0)
                product.revenue_30d = float(agg.revenue_30d or 0)
                product.sold_90d = int(agg.sold_90d or 0)
                product.revenue_90d = float(agg.revenue_90d or 0)
                product.sold_365d = int(agg.sold_365d or 0)
                product.revenue_365d = float(agg.revenue_365d or 0)
                product.sold_all_time = int(agg.sold_all or 0)
                product.revenue_all_time = float(agg.revenue_all or 0)
                product.total_sold = int(agg.sold_90d or 0)  # legacy field — keeps 90d
                product.total_revenue = float(agg.revenue_90d or 0)
                product.last_sold_at = agg.last_sold_at
            else:
                # Product appears in line items but all are cancelled/refunded → zero out
                product.sold_30d = 0
                product.revenue_30d = 0.0
                product.sold_90d = 0
                product.revenue_90d = 0.0
                product.sold_365d = 0
                product.revenue_365d = 0.0
                product.sold_all_time = 0
                product.revenue_all_time = 0.0
                product.total_sold = 0
                product.total_revenue = 0.0

            self._update_computed_fields(product)
            recomputed += 1

        self.db.commit()
        return recomputed

    def _reset_products_with_no_line_items(self):
        """
        Zero out sales counters for products that have NO line items in the table.
        Called after a full backfill — products absent from 365 days of orders
        should tip into the "obsolete" tier.
        """
        from sqlalchemy import select, exists, and_

        # Find products with stock but no line items in the table at all
        sub = select(OrderLineItem.shopify_product_id).where(
            and_(
                OrderLineItem.shopify_product_id == Product.shopify_id,
                OrderLineItem.is_cancelled.is_(False),
                OrderLineItem.is_refunded.is_(False),
            )
        )

        products = self.db.query(Product).filter(
            ~exists(sub),
            Product.shopify_id.isnot(None),
        ).all()

        zeroed = 0
        for product in products:
            if (product.sold_30d or 0) > 0 or (product.sold_90d or 0) > 0 or (product.sold_365d or 0) > 0:
                product.sold_30d = 0
                product.revenue_30d = 0.0
                product.sold_90d = 0
                product.revenue_90d = 0.0
                product.sold_365d = 0
                product.revenue_365d = 0.0
                # Don't touch sold_all_time / last_sold_at — those reflect lifetime, not the 365d window
                self._update_computed_fields(product)
                zeroed += 1

        if zeroed > 0:
            self.db.commit()
            print(f"[Inventory] Zeroed {zeroed} products with no recent line items")
        return zeroed

    def get_order_history_for_product(self, shopify_product_id: str, days: int = 365) -> List[Dict[str, Any]]:
        """
        Return raw order history for a single product. Useful for the per-product
        drill-down view (timeline of every sale, with order numbers and dates).
        """
        since = datetime.now(timezone.utc) - timedelta(days=days)
        items = self.db.query(OrderLineItem).filter(
            OrderLineItem.shopify_product_id == shopify_product_id,
            OrderLineItem.order_created_at >= since,
            OrderLineItem.is_cancelled.is_(False),
        ).order_by(OrderLineItem.order_created_at.desc()).all()

        return [
            {
                "line_item_id": li.id,
                "order_id": li.order_id,
                "order_name": li.order_name,
                "sku": li.sku,
                "quantity": li.current_quantity,
                "unit_price": float(li.unit_price) if li.unit_price else 0.0,
                "revenue": float(li.revenue) if li.revenue else 0.0,
                "is_refunded": li.is_refunded,
                "order_created_at": li.order_created_at.isoformat() if li.order_created_at else None,
            }
            for li in items
        ]

    def recompute_inventory_health(self) -> Dict[str, Any]:
        """
        Recompute stock_health, dead_stock_tier, velocity, demand, urgency on every product
        using the existing sold_30d / sold_90d / sold_365d buckets.

        Does NOT call Shopify — runs entirely on local data. Use this after fixing
        classification logic to refresh the dashboard without a full re-sync.
        """
        products = self.db.query(Product).filter(
            Product.inventory_status.isnot(None)
        ).all()

        recomputed = 0
        for product in products:
            self._update_computed_fields(product)
            recomputed += 1

        self.db.commit()
        cache.invalidate("inventory:dashboard")
        cache.invalidate("inventory:health")
        cache.invalidate("inventory:action_center")

        # Build tier breakdown for the response
        tier_counts = {"slow": 0, "stale": 0, "dead": 0, "obsolete": 0, "healthy": 0}
        for p in products:
            if p.dead_stock_tier in tier_counts:
                tier_counts[p.dead_stock_tier] += 1
            elif p.inventory_quantity and p.inventory_quantity > 0:
                tier_counts["healthy"] += 1

        return {
            "products_recomputed": recomputed,
            "tier_breakdown": tier_counts,
        }

    # =========================================================================
    # CO-PURCHASE / ANCHOR PRODUCT ANALYSIS
    # =========================================================================

    def sync_co_purchase_data(self, days: int = 90) -> Dict[str, Any]:
        """
        Fetch co-purchase patterns from Shopify orders and update products.
        Detects anchor products — products that consistently drive multi-item carts.
        """
        from app.services.shopify_service import ShopifyService
        shopify_svc = ShopifyService()

        co_data = shopify_svc.get_co_purchase_data(days=days)
        if not co_data:
            return {"error": "No co-purchase data found", "products_updated": 0}

        # Find max co_purchase_count for normalization
        max_co = max((v["co_purchase_count"] for v in co_data.values()), default=1)
        max_companions = max((v["avg_cart_companions"] for v in co_data.values()), default=1)

        updated = 0
        products = self.db.query(Product).all()
        product_map = {p.shopify_id: p for p in products}

        for shopify_id, stats in co_data.items():
            product = product_map.get(str(shopify_id))
            if not product:
                continue

            product.co_purchase_count = stats["co_purchase_count"]
            product.avg_cart_companions = stats["avg_cart_companions"]

            # Cart revenue multiplier: avg cart total / product price
            try:
                price = float(product.price) if product.price else 0.0
            except (ValueError, TypeError):
                price = 0.0

            if price > 0 and stats["avg_cart_total"] > 0:
                product.cart_revenue_multiplier = round(stats["avg_cart_total"] / price, 2)
            else:
                product.cart_revenue_multiplier = 1.0

            # Top 5 companions with product titles
            companion_list = []
            for comp_id, count in list(stats["companions"].items())[:5]:
                comp_product = product_map.get(str(comp_id))
                companion_list.append({
                    "shopify_id": comp_id,
                    "title": comp_product.title if comp_product else f"Product {comp_id}",
                    "count": count,
                })
            product.top_companions = companion_list

            # Anchor score (0-100): how much this product drives multi-product sales
            total_orders = stats["co_purchase_count"] + stats["solo_purchase_count"]
            if total_orders > 0:
                co_purchase_ratio = stats["co_purchase_count"] / total_orders  # % of orders that are multi-item
                frequency_norm = min(stats["co_purchase_count"] / max(max_co, 1), 1.0)
                companions_norm = min(stats["avg_cart_companions"] / max(max_companions, 1), 1.0)
                multiplier_norm = min((product.cart_revenue_multiplier - 1.0) / 4.0, 1.0)  # 5x multiplier = max

                anchor = (
                    co_purchase_ratio * 35 +       # 35%: how often bought with others
                    frequency_norm * 25 +           # 25%: absolute volume of multi-item orders
                    companions_norm * 20 +          # 20%: avg companions per cart
                    max(multiplier_norm, 0) * 20    # 20%: revenue amplification
                )
                product.anchor_score = min(int(round(anchor)), 100)
            else:
                product.anchor_score = 0

            # Recalculate demand score with the new anchor factor
            product.demand_score = self._calculate_demand_score(product)
            product.urgency_score = self._calculate_urgency_score(product)

            updated += 1

        self.db.commit()
        cache.invalidate("inventory:dashboard")
        cache.invalidate("inventory:action_center")

        # Summary stats
        anchor_products = self.db.query(Product).filter(
            Product.anchor_score >= 50
        ).count()

        print(f"[Inventory] Co-purchase sync: {updated} products updated, {anchor_products} anchor products detected")

        return {
            "products_updated": updated,
            "orders_analyzed": len(co_data),
            "anchor_products": anchor_products,
            "days_analyzed": days,
        }

    def get_anchor_products(self, limit: int = 50) -> List[Dict]:
        """Get products ranked by anchor score."""
        products = self.db.query(Product).filter(
            Product.anchor_score > 0
        ).order_by(Product.anchor_score.desc()).limit(limit).all()

        result = []
        for p in products:
            d = self._product_to_inventory_dict(p)
            d["anchor_score"] = p.anchor_score
            d["co_purchase_count"] = p.co_purchase_count
            d["avg_cart_companions"] = p.avg_cart_companions
            d["cart_revenue_multiplier"] = p.cart_revenue_multiplier
            d["top_companions"] = p.top_companions
            result.append(d)
        return result

    # =========================================================================
    # HELPERS
    # =========================================================================

    def _product_to_inventory_dict(self, p: Product) -> Dict:
        """Convert a Product to an inventory-focused dict."""
        # Days since last sale (computed from last_sold_at if available)
        days_since_last_sale = None
        if p.last_sold_at:
            try:
                delta = datetime.now(timezone.utc) - p.last_sold_at
                days_since_last_sale = delta.days
            except (TypeError, ValueError):
                days_since_last_sale = None

        return {
            "id": p.id,
            "shopify_id": p.shopify_id,
            "title": p.title,
            "sku": p.sku,
            "handle": p.handle,
            "product_type": p.product_type,
            "price": p.price,
            "inventory_quantity": p.inventory_quantity,
            "inventory_status": p.inventory_status,
            "inventory_velocity": p.inventory_velocity,
            "days_of_supply": p.days_of_supply,
            "demand_score": p.demand_score,
            "stock_health": p.stock_health,
            "dead_stock_tier": p.dead_stock_tier,
            "last_sold_at": p.last_sold_at.isoformat() if p.last_sold_at else None,
            "days_since_last_sale": days_since_last_sale,
            "low_stock_threshold": p.low_stock_threshold,
            "sold_30d": p.sold_30d,
            "revenue_30d": p.revenue_30d,
            "sold_90d": p.sold_90d,
            "revenue_90d": p.revenue_90d,
            "sold_365d": p.sold_365d,
            "revenue_365d": p.revenue_365d,
            "ga4_sessions": p.ga4_sessions,
            "gsc_impressions": p.gsc_impressions,
            "active_subscribers": p.active_subscribers,
            "last_stockout_date": p.last_stockout_date.isoformat() if p.last_stockout_date else None,
            "stockout_frequency_90d": p.stockout_frequency_90d,
            "urgency_score": p.urgency_score,
            "revenue_lost_est": p.revenue_lost_est,
            "suggested_reorder_qty": p.suggested_reorder_qty,
            "inventory_by_location": p.inventory_by_location,
            "anchor_score": p.anchor_score,
            "top_companions": p.top_companions,
            "last_inventory_sync": p.last_inventory_sync.isoformat() if p.last_inventory_sync else None,
        }

    def export_to_excel(self, view: str = "products", status: Optional[str] = None, product_ids: Optional[List[str]] = None) -> bytes:
        """
        Generate an Excel file with inventory data.
        view: "products" (all products) or "action_center" (grouped by action).
        Returns bytes of the .xlsx file.
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        # Style definitions
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            bottom=Side(style="thin", color="DDDDDD"),
        )
        number_fmt_int = '#,##0'
        number_fmt_money = '#,##0.00'
        number_fmt_dec = '#,##0.00'

        def style_header(ws, columns):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        def write_product_row(ws, row_idx, p):
            location_str = ""
            if p.inventory_by_location:
                parts = [f"{name}: {qty}" for name, qty in p.inventory_by_location.items()]
                location_str = " | ".join(parts)

            values = [
                p.title,
                p.sku,
                p.product_type,
                p.price,
                p.inventory_quantity or 0,
                location_str,
                p.inventory_status or "",
                round(p.inventory_velocity or 0, 2),
                round(p.days_of_supply or 0, 1) if (p.days_of_supply or 0) < 999 else "∞",
                p.demand_score or 0,
                p.anchor_score or 0,
                p.sold_30d or 0,
                p.revenue_30d or 0,
                p.sold_90d or 0,
                p.revenue_90d or 0,
                p.active_subscribers or 0,
                p.ga4_sessions or 0,
                p.gsc_impressions or 0,
                p.gsc_clicks or 0,
                p.stock_health or "",
                p.suggested_reorder_qty or 0,
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        columns = [
            "Producto", "SKU", "Tipo", "Precio",
            "Cantidad", "Por Sucursal", "Status",
            "Velocidad (u/día)", "Días de Stock", "Demanda", "Anchor",
            "Vendidos 30d", "Revenue 30d", "Vendidos 90d", "Revenue 90d",
            "Suscriptores", "Sesiones GA4", "Impressiones GSC", "Clicks GSC",
            "Salud", "Reorden Sugerido",
        ]

        if view == "action_center":
            action_data = self.get_action_center()
            sections = [
                ("Restock Ahora", action_data["restock_now"]),
                ("Ordenar Pronto", action_data["order_soon"]),
                ("Sin Movimiento", action_data["slow_movers"]),
                ("Productos Estrella", action_data["star_products"]),
            ]
            for idx, (section_name, items) in enumerate(sections):
                if idx == 0:
                    ws = wb.active
                    ws.title = section_name
                else:
                    ws = wb.create_sheet(title=section_name)

                style_header(ws, columns)
                product_ids = [item["id"] for item in items]
                products = self.db.query(Product).filter(Product.id.in_(product_ids)).all()
                prod_map = {p.id: p for p in products}

                for row_idx, item in enumerate(items, 2):
                    p = prod_map.get(item["id"])
                    if p:
                        write_product_row(ws, row_idx, p)

                # Auto-width
                for col_idx in range(1, len(columns) + 1):
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16
                ws.column_dimensions["A"].width = 50  # Producto
                ws.column_dimensions["B"].width = 18  # SKU
                ws.column_dimensions["F"].width = 35  # Por Sucursal

        else:
            ws = wb.active
            ws.title = "Inventario"
            style_header(ws, columns)

            if product_ids:
                query = self.db.query(Product).filter(Product.id.in_(product_ids))
            else:
                query = self.db.query(Product).filter(Product.inventory_status.isnot(None))
                if status:
                    query = query.filter(Product.inventory_status == status)
            query = query.order_by(Product.demand_score.desc())
            products = query.all()

            for row_idx, p in enumerate(products, 2):
                write_product_row(ws, row_idx, p)

            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16
            ws.column_dimensions["A"].width = 50
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["F"].width = 35

        # Freeze header row
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
